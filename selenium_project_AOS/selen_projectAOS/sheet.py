import openpyxl
from openpyxl import Workbook
import json
import random
import string

class AOS_Sheet:

    # define the values in XL file
    # in each row are the kind of the values and in each column the number of the test
    def __init__(self):
        self.workbook = Workbook()
        self.workbook=openpyxl.load_workbook("Tests_AOS.xlsx")
        self.sheet = self.workbook.active
        self.tests=["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]

    # print all the products details in JSON format
    def xl_to_json_products(self):
        tests={}
        products_numbers=[self.sheet["A2"].value,self.sheet["A6"].value,self.sheet["A10"].value]
        products_details={}
        # pass on the xl table and make dictionary of tests and into hit dictionary of products and their details
        for column in self.sheet.iter_cols(1,11,1,13,True):
             if column[0] is None or column[1] is None:
                continue
             else:
                 test_number=column[0]
                 x=1
                 for i in products_numbers:

                    detail={
                     "Category":column[x],
                     "Product ID": column[x+1],
                     "Quantity": column[x+2],
                     "Color": column[x+3]
                    }
                    x+=4
                    products_details[i]=detail
                 tests[test_number]=products_details

        # print the tests dictionary in json format
        print(json.dumps(tests,indent=4))

    # get test number and product number and return the match category from the xl table
    def get_Category(self,test_number,product_number):
        if product_number==1:
            category=self.sheet[f"{self.tests[test_number-1]+'2'}"].value
            return category
        if product_number==2:
            category=self.sheet[f"{self.tests[test_number-1]+'6'}"].value
            return category
        if product_number==3:
            category=self.sheet[f"{self.tests[test_number-1]+'10'}"].value
            return category

    # get test number and product number and return the match product id from the xl table
    def get_Product_ID(self,test_number,product_number):
        if product_number == 1:
            id = self.sheet[f"{self.tests[test_number-1] + '3'}"].value
            return id
        if product_number == 2:
            id = self.sheet[f"{self.tests[test_number-1] + '7'}"].value
            return id
        if product_number == 3:
            id = self.sheet[f"{self.tests[test_number-1] + '11'}"].value
            return id

    # get test number and product number and return the match quantity from the xl table
    def get_Quantity(self,test_number,product_number):
        if product_number == 1:
            quantity = self.sheet[f"{self.tests[test_number-1] + '4'}"].value
            return quantity
        if product_number == 2:
            quantity = self.sheet[f"{self.tests[test_number-1] + '8'}"].value
            return quantity
        if product_number == 3:
            quantity = self.sheet[f"{self.tests[test_number-1] + '12'}"].value
            return quantity

    # get test number and product number and return the match color from the xl table
    def get_Color(self,test_number,product_number):
        if product_number == 1:
            color = self.sheet[f"{self.tests[test_number-1] + '5'}"].value
            return color
        if product_number == 2:
            color = self.sheet[f"{self.tests[test_number-1] + '9'}"].value
            return color
        if product_number == 3:
            color = self.sheet[f"{self.tests[test_number-1] + '13'}"].value
            return color

    # get test number and add to the XL file V if it passed and X if it failed
    def add_pass_or_fail(self,test_number,boo):
        if boo:
            self.sheet[f"{self.tests[test_number-1] + '26'}"]="V"
        else:
            self.sheet[f"{self.tests[test_number - 1] + '26'}"] = "X"
        self.workbook.save("Tests_AOS.xlsx")

    # get test number return the username of exist user from xl table
    def get_exist_username(self,test_number):
        username = self.sheet[f"{self.tests[test_number - 1] + '14'}"].value
        return username

    # get test number return the password of exist user
    def get_exist_password(self,test_number):
        password = self.sheet[f"{self.tests[test_number - 1] + '15'}"].value
        return password

    # return the username of new user
    def get_new_username(self,test_number):
        self.generate_username(test_number)
        username = self.sheet[f"{self.tests[test_number - 1] + '16'}"].value
        return username

    # return the password of new user
    def get_new_password(self,test_number):
        self.generate_password(test_number)
        password = self.sheet[f"{self.tests[test_number - 1] + '18'}"].value
        return password

    # return the the mail of new user
    def get_new_mail(self,test_number):
        self.generate_email(test_number)
        mail = self.sheet[f"{self.tests[test_number - 1] + '17'}"].value
        return mail

    def get_new_password_to_confirm(self,test_number):
        password = self.sheet[f"{self.tests[test_number - 1] + '18'}"].value
        return password

    # return the safe pay username of specific user
    def get_new_safepay_username(self,test_number):
        username = self.sheet[f"{self.tests[test_number - 1] + '19'}"].value
        return username

    # return the safe pay password of specific user
    def get_new_safepay_password(self,test_number):
        password = self.sheet[f"{self.tests[test_number - 1] + '20'}"].value
        return password

    # return the card-number of specific user
    def get_card_number(self,test_number):
        card_number = self.sheet[f"{self.tests[test_number - 1] + '21'}"].value
        return card_number

    # return the CVV of specific user
    def get_CVV(self,test_number):
        cvv = self.sheet[f"{self.tests[test_number - 1] + '22'}"].value
        return cvv

    # return the holder_name of specific user
    def get_holder_name(self,test_number):
        holder_name = self.sheet[f"{self.tests[test_number - 1] + '25'}"].value
        return holder_name

    # return the MM of specific user
    def get_MM(self,test_number):
        MM = self.sheet[f"{self.tests[test_number - 1] + '23'}"].value
        return MM

    # return the YYYY of specific user
    def get_YYYY(self,test_number):
        YYYY = self.sheet[f"{self.tests[test_number - 1] + '24'}"].value
        return YYYY

    def generate_password(self,test_number):
        # define data
        lower = string.ascii_lowercase
        upper = string.ascii_uppercase
        num = string.digits
        symbols = string.punctuation
        # string.ascii_letters
        # combine the data
        all = lower + upper + num + symbols
        # use random
        temp = random.sample(all, 8)
        # create the password
        password = "".join(temp)
        # save the new password in the table
        self.sheet[f"{self.tests[test_number - 1] + '18'}"]=password
        self.workbook.save("Tests_AOS.xlsx")

    def generate_username(self,test_number):
        # define data
        lower = string.ascii_lowercase
        upper = string.ascii_uppercase
        num = string.digits
        # string.ascii_letters
        # combine the data
        all = lower + upper + num
        # use random
        temp = random.sample(all, 7)
        # create the username
        username = "".join(temp)
        # save username in table
        self.sheet[f"{self.tests[test_number - 1] + '16'}"]=username
        self.workbook.save("Tests_AOS.xlsx")

    def generate_email(self,test_number):
        # define data
        lower = string.ascii_lowercase
        upper = string.ascii_uppercase
        num = string.digits
        # string.ascii_letters
        # combine the data
        all = lower + upper + num
        # use random
        temp = random.sample(all, 5)
        # create the email
        email = "".join(temp)+"@gmail.com"
        # print the password
        self.sheet[f"{self.tests[test_number - 1] + '17'}"]=email
        self.workbook.save("Tests_AOS.xlsx")








if __name__=="__main__":
    x=AOS_Sheet()
    x.xl_to_json_products()
    x.get_Category(3,1)
    x.get_Quantity(3,1)
    x.get_Color(3,1)
    x.get_Product_ID(3,1)
    x.generate_password()